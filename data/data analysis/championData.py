import json
from openpyxl import Workbook
from openpyxl import load_workbook
import string
import os.path

version = "10.14.1"
json_src = "../dragontail/dragontail-" + version + "/" + version + "/data/en_US/champion.json"
xlsx_src = "Champion Stat.xlsx"
#change the file route when the game version was updated
with open(json_src, "r", encoding="utf-8") as f:
    champInfo = json.load(f)
    dic = json.dumps(champInfo)

#print(json.dumps(champInfo, indent="\t")) #for json file text

if os.path.isfile(xlsx_src):
    wb = load_workbook(filename=xlsx_src)
else:
    wb = Workbook()
output = wb.create_sheet(version)

champName = list(champInfo["data"].keys())
champStat = list(champInfo["data"]["Aatrox"]["stats"].keys())

#make index
letters = list(string.ascii_uppercase)
output["B2"] = "Id"
output["C2"] = "Name"

char_index = 3
for i in champStat:
    output[letters[char_index] + "2"] = i
    char_index += 1

#fill the contents
num_index = 3

for i in range(0, len(champName)):
    char_index = 3
    output["C" + str(num_index)] = champName[i]
    output["B" + str(num_index)] = champInfo["data"][champName[i]]["key"]
    for j in range(0, len(champStat)):
        output[letters[char_index] + str(num_index)] = champInfo["data"][champName[i]]["stats"][champStat[j]]
        char_index += 1
    num_index += 1

wb.save(xlsx_src)
"""
champName = ['Aatrox', 'Ahri', 'Akali', 'Alistar']
print(champInfo.items())
print(champInfo['data'].items())
print(champInfo['data']['Aatrox'].items())
print(champInfo['data']['Ahri']['stats'].items())
"""